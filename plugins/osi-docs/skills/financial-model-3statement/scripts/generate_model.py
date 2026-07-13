#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
financial-model-3statement | generate_model.py

月次60ヶ月の財務3表（PL/BS/CF）を「フル連動」で生成する。
連動の作法（AI OSI URI 事業計画モデルに準拠）:
  - PL は売上・原価を「ファネル / ドライバー階層」で積み上げる
  - BS は各勘定を独立に転がし、現金 =「前月現金 + 他の全BS勘定の増減」で算出
    → 資産 = 負債 + 純資産 が構造上ゼロ差で成立（Balance Check = 0）
  - CF は BS の差分から独立に再構築し、CF期末現金 = BS現金 を検算（連動Check = 0）
  - 消費税(仮払/仮受)と売掛金の入金サイトを実装
  - 立ち上がりはコンサバ（採用→稼働まで成熟ラグ × 成熟稼働率）

使い方:
  python3 generate_model.py --config config.json --out model.xlsx
  python3 generate_model.py --out model.xlsx          # 同梱DEFAULT_CONFIGで生成

生成後は必ず:
  1) recalc（xlsx スキルの scripts/recalc.py）で数式を再計算しエラー0を確認
  2) verify.py で Balance Check / 連動Check が全月0、必要資金(トラフ)を確認

注意: 本テンプレートの売上ドライバーは「人材紹介 × スクール」型
(稼働CA→送客→成約→紹介売上 / 受講料売上)。他業種に転用する場合は
PL のドライバー行ブロックだけ差し替える。BS/CF の連動機構は業種非依存でそのまま使える。
references/method.md を参照。
"""
import argparse, json, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as GL
from openpyxl.chart import BarChart, LineChart, Reference

NM = 60  # months

DEFAULT_CONFIG = {
    "title": "事業計画｜財務モデル（新会社）",
    "unit_economics": {
        "月間送客数 / CA（人）": 30,
        "送客→成約 成約率": 0.0833,
        "平均紹介料（円/件）": 1500000,
        "クロスセル受講率（成約比）": 0.8,
        "受講料（円/人）": 300000,
    },
    "ramp": {"採用→稼働 成熟ラグ（月）": 3, "成熟稼働率": 0.9},
    "cost": {
        "FCロイヤリティ（円/月/CA）": 500000,
        "CA人件費（円/年/CA）": 7000000,
        "スクール運営費（円/受講者）": 250000,
        "FC加盟金（初期一括, 円）": 6000000,
    },
    "wc_tax_cap": {
        "売上回収サイト（月）": 1,
        "法人実効税率": 0.35,
        "消費税率": 0.10,
        "創業時資本金（円）": 1000000,
        "増資のうち資本金比率": 0.5,
    },
    # 販管費: 科目ごとに FY1..FY5 の年額（円）
    "sga_fy": {
        "役員報酬": [24e6, 30e6, 36e6, 42e6, 48e6],
        "管理・開拓人件費": [15e6, 35e6, 70e6, 100e6, 140e6],
        "法定福利費": [6e6, 10e6, 18e6, 24e6, 30e6],
        "士業・外注費": [6e6, 8e6, 10e6, 14e6, 18e6],
        "広告宣伝費": [8e6, 12e6, 14e6, 18e6, 24e6],
        "採用教育費": [4e6, 5e6, 8e6, 10e6, 12e6],
        "通信費（AI API/SaaS）": [4e6, 6e6, 8e6, 12e6, 14e6],
        "支払手数料": [2e6, 3e6, 4e6, 6e6, 8e6],
        "地代家賃": [0, 0, 0, 2e6, 4e6],
        "消耗品費": [1e6, 1e6, 2e6, 2e6, 2e6],
        "雑費": [0, 0, 0, 0, 0],
    },
    # 課税仕入(消費税の仕入控除)に含める販管費科目（給与・法定福利などは非課税のため除く）
    "taxable_sga": ["士業・外注費", "広告宣伝費", "採用教育費", "通信費（AI API/SaaS）",
                     "支払手数料", "地代家賃", "消耗品費"],
    # 月末CA人数（60ヶ月）。コンサバな採用カーブを入れる
    "headcount": [1, 1, 1, 1, 2, 2, 2, 2, 3, 3, 3, 3,
                  4, 4, 5, 5, 6, 6, 7, 7, 7, 8, 8, 8,
                  9, 9, 10, 10, 11, 11, 12, 12, 13, 13, 14, 14,
                  15, 15, 16, 16, 17, 17, 18, 18, 19, 19, 20, 20,
                  21, 21, 22, 22, 23, 23, 24, 24, 25, 25, 26, 26],
    # 増資（円・60ヶ月）。0で埋め、必要月に金額を入れる。空ならverify後に必要資金を見て設定する
    "raise": [0] * NM,
}

YEN='#,##0;(#,##0);-'; PCT='0.0%'; NUM1='#,##0.0'
BLUE=Font(name='Arial',color='0000FF'); BLACK=Font(name='Arial',color='000000')
GREEN=Font(name='Arial',color='008000'); BOLD=Font(name='Arial',bold=True)
BOLDW=Font(name='Arial',bold=True,color='FFFFFF')
HDR=PatternFill('solid',fgColor='1F3864'); SEC=PatternFill('solid',fgColor='D9E1F2')
TOT=PatternFill('solid',fgColor='EDEDED'); CHK=PatternFill('solid',fgColor='FFF2CC')


def put(ws,r,c,v,font=BLACK,fmt=None,fill=None,al=None):
    cell=ws.cell(row=r,column=c,value=v); cell.font=font
    if fmt: cell.number_format=fmt
    if fill: cell.fill=fill
    if al: cell.alignment=Alignment(horizontal=al)
    return cell


def build(cfg, out_path):
    ue=cfg["unit_economics"]; rmp=cfg["ramp"]; cost=cfg["cost"]; wtc=cfg["wc_tax_cap"]
    sga_fy=cfg["sga_fy"]; taxable=cfg["taxable_sga"]; hc=cfg["headcount"]; raise_=cfg["raise"]
    assert len(hc)==NM and len(raise_)==NM, "headcount / raise は60要素必要"
    wb=Workbook()

    # ===== 前提・KPI =====
    a=wb.active; a.title='前提・KPI'
    a.column_dimensions['A'].width=34; a.column_dimensions['B'].width=14
    for i in range(NM): a.column_dimensions[GL(3+i)].width=8
    put(a,1,1,cfg["title"]+'｜前提・KPI',BOLD)
    put(a,2,1,'青字=入力値。月次60ヶ月の3表(PL/BS/CF)がこのシートを関数参照。',Font(name='Arial',italic=True,size=9))
    PR={}; rr=4
    def sec(t):
        nonlocal rr
        put(a,rr,1,t,BOLDW,fill=HDR); a.cell(row=rr,column=2).fill=HDR; rr+=1
    def kpi(label,val,fmt=NUM1):
        nonlocal rr
        put(a,rr,1,label); put(a,rr,2,val,BLUE,fmt); PR[label]=rr; rr+=1
    sec('■ 売上ファネル（稼働CA起点）')
    for k in ["月間送客数 / CA（人）","送客→成約 成約率","平均紹介料（円/件）","クロスセル受講率（成約比）","受講料（円/人）"]:
        kpi(k,ue[k], PCT if ('率' in k) else (YEN if '円' in k else NUM1))
    sec('■ 立ち上がり（コンサバ）')
    kpi('採用→稼働 成熟ラグ（月）',rmp['採用→稼働 成熟ラグ（月）'])
    kpi('成熟稼働率',rmp['成熟稼働率'],PCT)
    sec('■ コスト単価')
    for k in ["FCロイヤリティ（円/月/CA）","CA人件費（円/年/CA）","スクール運営費（円/受講者）","FC加盟金（初期一括, 円）"]:
        kpi(k,cost[k],YEN)
    sec('■ 運転資本・税・資本')
    kpi('売上回収サイト（月）',wtc['売上回収サイト（月）'])
    kpi('法人実効税率',wtc['法人実効税率'],PCT)
    kpi('消費税率',wtc['消費税率'],PCT)
    kpi('創業時資本金（円）',wtc['創業時資本金（円）'],YEN)
    kpi('増資のうち資本金比率',wtc['増資のうち資本金比率'],PCT)
    rr+=1
    put(a,rr,1,'■ 販管費 内訳（円, FY別）',BOLDW,fill=HDR)
    for c in range(2,7): a.cell(row=rr,column=c).fill=HDR
    rr+=1
    put(a,rr,1,'年度',BOLD)
    for i in range(5): put(a,rr,2+i,f'FY{i+1}',BOLD,al='center')
    rr+=1
    SGA={}
    for label,vals in sga_fy.items():
        put(a,rr,1,label)
        for i,v in enumerate(vals): put(a,rr,2+i,v,BLUE,YEN)
        SGA[label]=rr; rr+=1
    rr+=1
    put(a,rr,1,'■ 月次スケジュール（CA人数・増資）',BOLDW,fill=HDR)
    for c in range(3,3+NM): a.cell(row=rr,column=c).fill=HDR
    rr+=1
    put(a,rr,1,'月番号',BOLD)
    for i in range(NM): put(a,rr,3+i,i+1,Font(name='Arial',size=8,bold=True),al='center')
    rr+=1
    HC_ROW=rr; put(a,rr,1,'CA人数（月末・実数）')
    for i in range(NM): put(a,rr,3+i,hc[i],BLUE,NUM1)
    rr+=1
    RAISE_ROW=rr; put(a,rr,1,'増資（円）')
    for i in range(NM): put(a,rr,3+i,raise_[i],BLUE,YEN)

    P="'前提・KPI'!"
    def pc(label): return f"{P}$B${PR[label]}"
    HCR=f"{P}$C${HC_ROW}:${GL(2+NM)}${HC_ROW}"
    def RAISEc(i): return f"{P}{GL(3+i)}${RAISE_ROW}"
    def sgaFY(label): return f"{P}$B${SGA[label]}:$F${SGA[label]}"

    # ===== 月次三表 =====
    m=wb.create_sheet('月次三表（FY1-5）')
    m.column_dimensions['A'].width=32; m.column_dimensions['B'].width=13
    for i in range(NM): m.column_dimensions[GL(3+i)].width=11
    put(m,1,1,cfg["title"]+'｜月次三表（60ヶ月）PL→BS→CF フル連動',BOLD)
    put(m,2,1,'単位:円。緑=他シート参照/黒=計算。B列=期初(基準値)。',Font(name='Arial',italic=True,size=9))
    def col(i): return GL(3+i)
    def prev(i): return 'B' if i==0 else GL(3+i-1)
    R={}; rr2=3
    def row(key,label,ind=0,bold=False,fill=None):
        nonlocal rr2
        R[key]=rr2; put(m,rr2,1,('　'*ind)+label,BOLD if bold else BLACK)
        if fill:
            for c in [1,2]+[3+i for i in range(NM)]: m.cell(row=rr2,column=c).fill=fill
        rr2+=1
    row('mnum','月番号',bold=True,fill=SEC); row('fy','年度(FY)',bold=True)
    row('is','■ 損益計算書（PL）',bold=True,fill=SEC)
    row('sales','売上高',bold=True)
    row('rev_ref','紹介売上',1); row('deals','成約数（件）',2); row('sokyaku','送客数（人）',3)
    row('act','稼働CA換算',4); row('rate','成約率',3); row('fee','平均紹介料（円/件）',2)
    row('rev_sch','受講料売上',1); row('stu','受講者数（人）',2); row('fees','受講料（円/人）',2)
    row('cogs','売上原価',bold=True)
    row('c_roy','FCロイヤリティ',1); row('hc','CA人数（月末）',2)
    row('c_ca','CA人件費',1); row('c_sch','スクール運営費',1)
    row('gp','売上総利益',bold=True)
    row('sga','販管費',bold=True)
    SGA_KEYS=[]
    for k in SGA:
        kk='s_'+str(SGA[k]); row(kk,k,1); SGA_KEYS.append((kk,k))
    row('kamei','FC加盟金',1); row('taxbuy','（補助）課税仕入額',1)
    row('op','営業利益',bold=True); row('noi','営業外損益',1); row('ord','経常利益',bold=True)
    row('pbt','税引前当期純利益',bold=True); row('tax','法人税等',1); row('ni','当期純利益',bold=True)
    row('bs','■ 貸借対照表（BS）',bold=True,fill=SEC)
    row('assets','資産',bold=True); row('ca_cur','流動資産',1)
    row('cash','現金及び預金',2); row('ar','売掛金',2)
    row('ar_new','当月売上額',3); row('ar_col','売掛金消し込み額',3)
    row('vin','仮払消費税',2); row('fa','固定資産',1)
    row('liab','負債',bold=True); row('ap','買掛金',1); row('taxp','未払法人税等',1); row('vout','仮受消費税',1)
    row('equity','純資産',bold=True); row('cap','資本金',1); row('capr','資本準備金',1); row('ret','繰越利益剰余金',1)
    row('tle','負債・純資産合計',bold=True); row('bchk','Balance Check（資産−負債−純資産=0）',bold=True,fill=CHK)
    row('cf','■ キャッシュフロー計算書（CF・独立）',bold=True,fill=SEC)
    row('o_ni','当期純利益',1); row('o_ar','売掛金の増減（増△）',1); row('o_vin','仮払消費税の増減（増△）',1)
    row('o_ap','買掛金の増減（増＋）',1); row('o_taxp','未払法人税の増減（増＋）',1); row('o_vout','仮受消費税の増減（増＋）',1)
    row('ocf','営業CF',bold=True); row('icf','投資CF',bold=True); row('fin','増資',1); row('fcf','財務CF',bold=True)
    row('chg','当月キャッシュ増減',bold=True); row('beg','現金期首',1); row('end','現金期末',bold=True)
    row('tie','連動Check（CF期末−BS現金=0）',bold=True,fill=CHK)
    row('nf','■ 資金繰り（無調達ベース）',bold=True,fill=SEC); row('nfcash','現金残高（増資ゼロ想定）',bold=True)

    def C(i,k): return f"{col(i)}{R[k]}"
    def Cp(i,k): return f"{prev(i)}{R[k]}"
    def setv(key,fn,fmt=YEN,font=BLACK,base=None):
        r=R[key]
        if base is not None:
            bc=m.cell(row=r,column=2,value=base); bc.number_format=fmt; bc.font=BLUE
        for i in range(NM):
            c=m.cell(row=r,column=3+i,value=fn(i)); c.number_format=fmt; c.font=font
    FYstart=lambda i:f"MOD({C(i,'mnum')}-1,12)=0"

    setv('mnum',lambda i:i+1,NUM1,BOLD)
    setv('fy',lambda i:f"=INT(({C(i,'mnum')}-1)/12)+1",NUM1)
    lag=pc('採用→稼働 成熟ラグ（月）'); rate=pc('成熟稼働率')
    setv('act',lambda i:f"=IF({C(i,'mnum')}>{lag},INDEX({HCR},1,{C(i,'mnum')}-{lag})*{rate},0)",NUM1,GREEN)
    setv('sokyaku',lambda i:f"={C(i,'act')}*{pc('月間送客数 / CA（人）')}",NUM1,GREEN)
    setv('rate',lambda i:f"={pc('送客→成約 成約率')}",PCT,GREEN)
    setv('deals',lambda i:f"={C(i,'sokyaku')}*{C(i,'rate')}",NUM1)
    setv('fee',lambda i:f"={pc('平均紹介料（円/件）')}",YEN,GREEN)
    setv('rev_ref',lambda i:f"={C(i,'deals')}*{C(i,'fee')}")
    setv('stu',lambda i:f"={C(i,'deals')}*{pc('クロスセル受講率（成約比）')}",NUM1,GREEN)
    setv('fees',lambda i:f"={pc('受講料（円/人）')}",YEN,GREEN)
    setv('rev_sch',lambda i:f"={C(i,'stu')}*{C(i,'fees')}")
    setv('sales',lambda i:f"={C(i,'rev_ref')}+{C(i,'rev_sch')}",font=BOLD)
    setv('hc',lambda i:f"=INDEX({HCR},1,{C(i,'mnum')})",NUM1,GREEN)
    setv('c_roy',lambda i:f"={C(i,'hc')}*{pc('FCロイヤリティ（円/月/CA）')}")
    setv('c_ca',lambda i:f"={C(i,'hc')}*{pc('CA人件費（円/年/CA）')}/12")
    setv('c_sch',lambda i:f"={C(i,'stu')}*{pc('スクール運営費（円/受講者）')}")
    setv('cogs',lambda i:f"={C(i,'c_roy')}+{C(i,'c_ca')}+{C(i,'c_sch')}",font=BOLD)
    setv('gp',lambda i:f"={C(i,'sales')}-{C(i,'cogs')}",font=BOLD)
    for kk,label in SGA_KEYS:
        setv(kk,(lambda lab:(lambda i:f"=INDEX({sgaFY(lab)},1,{C(i,'fy')})/12"))(label),YEN,GREEN)
    setv('kamei',lambda i:(f"={pc('FC加盟金（初期一括, 円）')}" if i==0 else "0"))
    setv('sga',lambda i:"="+"+".join(C(i,kk) for kk,_ in SGA_KEYS)+f"+{C(i,'kamei')}",font=BOLD)
    tax_keys=[('s_'+str(SGA[k])) for k in taxable if k in SGA]
    setv('taxbuy',lambda i:f"={C(i,'c_roy')}+{C(i,'c_sch')}"+("".join('+'+C(i,kk) for kk in tax_keys)))
    setv('op',lambda i:f"={C(i,'gp')}-{C(i,'sga')}",font=BOLD)
    setv('noi',lambda i:"0"); setv('ord',lambda i:f"={C(i,'op')}+{C(i,'noi')}",font=BOLD)
    setv('pbt',lambda i:f"={C(i,'ord')}",font=BOLD)
    setv('tax',lambda i:f"=IF({C(i,'pbt')}>0,{C(i,'pbt')}*{pc('法人実効税率')},0)")
    setv('ni',lambda i:f"={C(i,'pbt')}-{C(i,'tax')}",font=BOLD)
    # BS rolls
    setv('ar_new',lambda i:f"={C(i,'sales')}")
    isite=pc('売上回収サイト（月）')
    setv('ar_col',lambda i:f"=IF({C(i,'mnum')}>{isite},INDEX($C${R['sales']}:${GL(2+NM)}${R['sales']},1,{C(i,'mnum')}-{isite}),0)")
    setv('ar',lambda i:f"={Cp(i,'ar')}+{C(i,'ar_new')}-{C(i,'ar_col')}",base=0)
    vr=pc('消費税率')
    setv('vin',lambda i:f"=IF({FYstart(i)},{C(i,'taxbuy')}*{vr},{Cp(i,'vin')}+{C(i,'taxbuy')}*{vr})",base=0)
    setv('vout',lambda i:f"=IF({FYstart(i)},{C(i,'sales')}*{vr},{Cp(i,'vout')}+{C(i,'sales')}*{vr})",base=0)
    setv('ap',lambda i:f"={C(i,'c_roy')}+{C(i,'c_sch')}")
    setv('taxp',lambda i:f"=IF({FYstart(i)},{C(i,'tax')},{Cp(i,'taxp')}+{C(i,'tax')})",base=0)
    setv('cap',lambda i:f"={Cp(i,'cap')}+{RAISEc(i)}*{pc('増資のうち資本金比率')}",base=wtc['創業時資本金（円）'])
    setv('capr',lambda i:f"={Cp(i,'capr')}+{RAISEc(i)}*(1-{pc('増資のうち資本金比率')})",base=0)
    setv('ret',lambda i:f"={Cp(i,'ret')}+{C(i,'ni')}",base=0)
    setv('fa',lambda i:"0")
    def cash_f(i):
        d=lambda k:f"({C(i,k)}-{Cp(i,k)})"
        return (f"={Cp(i,'cash')}-{d('ar')}-{d('vin')}+{d('ap')}+{d('taxp')}+{d('vout')}"
                f"+{d('cap')}+{d('capr')}+{d('ret')}")
    setv('cash',cash_f,base=wtc['創業時資本金（円）'])
    setv('ca_cur',lambda i:f"={C(i,'cash')}+{C(i,'ar')}+{C(i,'vin')}",font=BOLD)
    setv('assets',lambda i:f"={C(i,'ca_cur')}+{C(i,'fa')}",font=BOLD)
    setv('liab',lambda i:f"={C(i,'ap')}+{C(i,'taxp')}+{C(i,'vout')}",font=BOLD)
    setv('equity',lambda i:f"={C(i,'cap')}+{C(i,'capr')}+{C(i,'ret')}",font=BOLD)
    setv('tle',lambda i:f"={C(i,'liab')}+{C(i,'equity')}",font=BOLD)
    setv('bchk',lambda i:f"={C(i,'assets')}-{C(i,'liab')}-{C(i,'equity')}",font=BOLD)
    # CF independent
    setv('o_ni',lambda i:f"={C(i,'ni')}")
    setv('o_ar',lambda i:f"=-({C(i,'ar')}-{Cp(i,'ar')})")
    setv('o_vin',lambda i:f"=-({C(i,'vin')}-{Cp(i,'vin')})")
    setv('o_ap',lambda i:f"={C(i,'ap')}-{Cp(i,'ap')}")
    setv('o_taxp',lambda i:f"={C(i,'taxp')}-{Cp(i,'taxp')}")
    setv('o_vout',lambda i:f"={C(i,'vout')}-{Cp(i,'vout')}")
    setv('ocf',lambda i:f"={C(i,'o_ni')}+{C(i,'o_ar')}+{C(i,'o_vin')}+{C(i,'o_ap')}+{C(i,'o_taxp')}+{C(i,'o_vout')}",font=BOLD)
    setv('icf',lambda i:"0",font=BOLD)
    setv('fin',lambda i:f"=({C(i,'cap')}-{Cp(i,'cap')})+({C(i,'capr')}-{Cp(i,'capr')})")
    setv('fcf',lambda i:f"={C(i,'fin')}",font=BOLD)
    setv('chg',lambda i:f"={C(i,'ocf')}+{C(i,'icf')}+{C(i,'fcf')}",font=BOLD)
    setv('beg',lambda i:f"={Cp(i,'cash')}")
    setv('end',lambda i:f"={C(i,'beg')}+{C(i,'chg')}",font=BOLD)
    setv('tie',lambda i:f"={C(i,'end')}-{C(i,'cash')}",font=BOLD)
    setv('nfcash',lambda i:(f"={pc('創業時資本金（円）')}+{C(i,'ocf')}+{C(i,'icf')}" if i==0
                            else f"={Cp(i,'nfcash')}+{C(i,'ocf')}+{C(i,'icf')}"),font=BOLD)

    MR={k:R[k] for k in R}
    MS="'月次三表（FY1-5）'!"
    def msum(key,fy): r=MR[key]; return f"=SUM({MS}{GL(3+fy*12)}{r}:{GL(3+fy*12+11)}{r})"
    def mend(key,fy): r=MR[key]; return f"={MS}{GL(3+fy*12+11)}{r}"

    # ===== 財務三表（年次・参照） =====
    y=wb.create_sheet('財務三表（年次・参照）')
    y.column_dimensions['A'].width=34
    for c in 'BCDEF': y.column_dimensions[c].width=16
    put(y,1,1,'財務三表（年次）＝月次の集計参照（PL/CF=合計, BS=期末）',BOLD)
    put(y,2,1,'単位:円。月次三表シートを関数参照（連動）。',Font(name='Arial',italic=True,size=9))
    for i in range(5): put(y,3,2+i,f'FY{i+1}',BOLDW,al='center',fill=HDR)
    y.cell(row=3,column=1).fill=HDR
    YR={}; rr3=4
    def L(key,label,bold=False,fill=None,fmt=YEN,ind=0):
        nonlocal rr3
        YR[key]=rr3; put(y,rr3,1,('　'*ind)+label,BOLD if bold else BLACK)
        if fill:
            for c in range(1,7): y.cell(row=rr3,column=c).fill=fill
        rr3+=1
    def fl(key,mkey,bold=False):
        r=YR[key]
        for fy in range(5):
            c=y.cell(row=r,column=2+fy,value=msum(mkey,fy)); c.number_format=YEN; c.font=BOLD if bold else GREEN
    def fe(key,mkey,bold=False):
        r=YR[key]
        for fy in range(5):
            c=y.cell(row=r,column=2+fy,value=mend(mkey,fy)); c.number_format=YEN; c.font=BOLD if bold else GREEN
    def fc(key,fn,fmt=YEN,bold=False):
        r=YR[key]
        for fy in range(5):
            c=y.cell(row=r,column=2+fy,value=fn(fy,GL(2+fy))); c.number_format=fmt; c.font=BOLD if bold else BLACK
    L('is','■ 損益計算書（PL）',bold=True,fill=SEC)
    L('ca','CA人数（年平均）',fmt=NUM1,ind=1)
    L('sales','売上高',bold=True); L('rev_ref','紹介売上',ind=1); L('rev_sch','受講料売上',ind=1)
    L('cogs','売上原価',bold=True); L('gp','売上総利益',bold=True,fill=TOT); L('sga','販管費',ind=1)
    L('op','営業利益',bold=True,fill=TOT); L('opm','営業利益率',fmt=PCT,ind=1)
    L('tax','法人税等',ind=1); L('ni','当期純利益',bold=True,fill=TOT)
    L('bs','■ 貸借対照表（BS, 期末）',bold=True,fill=SEC)
    L('cash','現金及び預金',ind=1); L('ar','売掛金',ind=1); L('vin','仮払消費税',ind=1); L('ta','資産',bold=True)
    L('ap','買掛金',ind=1); L('taxp','未払法人税等',ind=1); L('vout','仮受消費税',ind=1); L('tl','負債',bold=True)
    L('te','純資産',bold=True); L('tle','負債・純資産合計',bold=True); L('bchk','Balance Check（=0）',bold=True,fill=CHK)
    L('cf','■ キャッシュフロー（CF）',bold=True,fill=SEC)
    L('ocf','営業CF',bold=True); L('icf','投資CF',ind=1); L('fcf','財務CF（増資）',ind=1)
    L('chg','キャッシュ増減',bold=True); L('endc','現金期末残高',bold=True,fill=TOT)
    fc('ca',lambda fy,c:f"=AVERAGE('前提・KPI'!{GL(3+fy*12)}{HC_ROW}:{GL(3+fy*12+11)}{HC_ROW})",NUM1)
    for fyc in range(5): y.cell(row=YR['ca'],column=2+fyc).font=GREEN
    fl('sales','sales',True); fl('rev_ref','rev_ref'); fl('rev_sch','rev_sch')
    fl('cogs','cogs',True); fl('gp','gp',True); fl('sga','sga'); fl('op','op',True)
    fc('opm',lambda fy,c:f"=IF({c}{YR['sales']}=0,0,{c}{YR['op']}/{c}{YR['sales']})",PCT)
    fl('tax','tax'); fl('ni','ni',True)
    fe('cash','cash'); fe('ar','ar'); fe('vin','vin'); fe('ta','assets',True)
    fe('ap','ap'); fe('taxp','taxp'); fe('vout','vout'); fe('tl','liab',True)
    fe('te','equity',True); fe('tle','tle',True)
    fc('bchk',lambda fy,c:f"={c}{YR['ta']}-{c}{YR['tle']}",YEN,True)
    fl('ocf','ocf',True); fl('icf','icf'); fl('fcf','fcf'); fl('chg','chg',True); fe('endc','end',True)

    # ===== サマリー =====
    s=wb.create_sheet('サマリー',0)
    s.column_dimensions['A'].width=30
    for c in 'BCDEF': s.column_dimensions[c].width=15
    A2="'財務三表（年次・参照）'!"
    put(s,1,1,cfg["title"]+'｜サマリー（コンサバ・フル3表連動）',BOLD)
    put(s,2,1,'単位:円。月次60ヶ月でPL→BS→CFを連動（BSで現金自閉＋CF reconcile）。',Font(name='Arial',italic=True,size=9))
    for i in range(5): put(s,4,2+i,f'FY{i+1}',BOLDW,al='center',fill=HDR)
    s.cell(row=4,column=1).fill=HDR
    rows=[('CA人数（年平均）','ca',NUM1),('売上高','sales',YEN),('売上総利益','gp',YEN),
          ('営業利益','op',YEN),('営業利益率','opm',PCT),('当期純利益','ni',YEN),('現金期末残高','endc',YEN)]
    r=5
    for disp,key,fmt in rows:
        put(s,r,1,disp,BOLD if disp in('売上高','営業利益') else BLACK)
        for fy in range(5):
            c=s.cell(row=r,column=2+fy,value=f"={A2}{GL(2+fy)}{YR[key]}"); c.font=GREEN; c.number_format=fmt
        r+=1
    nf=R['nfcash']; endr=R['end']
    ft=r+1
    put(s,ft,1,'■ 資金需要（無調達ベース）',BOLDW,fill=HDR)
    for c in range(2,4): s.cell(row=ft,column=c).fill=HDR
    def kv(o,label,f,font=GREEN):
        put(s,ft+o,1,label); c=s.cell(row=ft+o,column=3,value=f); c.number_format=YEN; c.font=font
    kv(1,'無調達ベース 最低現金（トラフ）',f"=MIN({MS}C{nf}:{GL(2+60)}{nf})")
    kv(2,'必要資金（トラフ補填）',f"=MAX(0,-C{ft+1})",BLACK)
    kv(3,'推奨調達額（＋20%バッファ）',f"=CEILING(C{ft+2}*1.2,10000000)",BLACK)
    kv(4,'実際の調達額（入力）',f"=SUM('前提・KPI'!C{RAISE_ROW}:{GL(2+60)}{RAISE_ROW})")
    kv(5,'調達後 最低現金',f"=MIN({MS}C{endr}:{GL(2+60)}{endr})")
    ct=ft+8
    put(s,ct,1,'（グラフ用）',Font(name='Arial',size=8,italic=True)); put(s,ct+1,1,'年度')
    for fy in range(5): put(s,ct+1,2+fy,f'FY{fy+1}',Font(name='Arial',size=9),al='center')
    put(s,ct+2,1,'売上高'); put(s,ct+3,1,'営業利益')
    for fy in range(5):
        s.cell(row=ct+2,column=2+fy,value=f"={A2}{GL(2+fy)}{YR['sales']}").number_format=YEN
        s.cell(row=ct+3,column=2+fy,value=f"={A2}{GL(2+fy)}{YR['op']}").number_format=YEN
    bar=BarChart(); bar.title='売上高 と 営業利益（5カ年・コンサバ）'; bar.height=8; bar.width=18
    bar.add_data(Reference(s,min_col=1,min_row=ct+2,max_col=6,max_row=ct+2),titles_from_data=True,from_rows=True)
    bar.set_categories(Reference(s,min_col=2,min_row=ct+1,max_col=6,max_row=ct+1))
    ln=LineChart(); ln.add_data(Reference(s,min_col=1,min_row=ct+3,max_col=6,max_row=ct+3),titles_from_data=True,from_rows=True)
    bar+=ln; s.add_chart(bar,'A'+str(ct+5))

    # ===== パイプライン =====
    p=wb.create_sheet('パイプライン明細')
    for j,h in enumerate(['#','獲得月','顧客','区分','内容','想定単価（円）','係数','想定売上（円）','ステータス','獲得元'],1):
        put(p,1,j,h,BOLDW,fill=HDR)
    p.column_dimensions['A'].width=5
    for c,w in zip('BCDEFGHIJ',[12,22,14,16,14,9,16,12,12]): p.column_dimensions[c].width=w
    put(p,2,1,'※ 実データ入力欄。想定売上＝想定単価×係数。',Font(name='Arial',italic=True,size=9))

    order=['サマリー','前提・KPI','月次三表（FY1-5）','財務三表（年次・参照）','パイプライン明細']
    wb._sheets.sort(key=lambda ws: order.index(ws.title) if ws.title in order else 99)
    wb.save(out_path)
    return out_path


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--config',default=None,help='config.json path (省略時は同梱DEFAULT_CONFIG)')
    ap.add_argument('--out',default='financial_model.xlsx')
    args=ap.parse_args()
    cfg=DEFAULT_CONFIG
    if args.config:
        with open(args.config,encoding='utf-8') as f: cfg=json.load(f)
    out=build(cfg,args.out)
    print('saved:',out)
    print('NEXT: 1) recalc  2) verify.py  でエラー0・両Check0・必要資金を確認すること')


if __name__=='__main__':
    main()
