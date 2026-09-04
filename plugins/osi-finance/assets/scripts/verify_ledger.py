#!/usr/bin/env python3
"""OSI Finance 台帳の検算（静かに間違う箇所に、落ちる検査を置く）

使い方:
    python3 verify_ledger.py <台帳フォルダ>
    python3 verify_ledger.py <台帳フォルダ> --months 2026-04 2026-05 2026-06 2026-07
    python3 verify_ledger.py <台帳フォルダ> --json      # 機械可読で出す

終了コード: 0=ERROR なし / 1=ERROR あり（記帳を止める）

なぜこれが要るか
----------------
台帳の事故は「静かに間違った答えが出る」形で起きる。借貸は一致したまま二重計上できるし、
税基準がズレていても計算は通り、それらしい差額が出る。人がルールを読んで気をつける方式は
守り忘れるが、exit 1 は忘れない。記帳のたびに走らせること。
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が必要です: pip3 install openpyxl")

JOURNAL = "仕訳台帳.xlsx"
AR = "請求管理台帳.xlsx"


def num(v):
    if v is None or v == "":
        return 0
    try:
        return float(str(v).replace(",", "").replace("¥", "").strip())
    except ValueError:
        return 0


def read_tab(path, tab, header_row=1):
    """ヘッダ名で解決して dict の list を返す。__row に実行番号を入れる。"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if tab not in wb.sheetnames:
        raise KeyError(f"{path.name} に「{tab}」タブがありません")
    ws = wb[tab]
    rows, headers = [], None
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < header_row:
            continue
        if headers is None:
            headers = [str(c).strip() if c is not None else "" for c in r]
            continue
        d = {h: v for h, v in zip(headers, r) if h}
        d["__row"] = i
        rows.append(d)
    wb.close()
    return headers or [], rows


class Report:
    def __init__(self):
        self.items = []

    def add(self, level, check, msg, detail=None):
        self.items.append({"level": level, "check": check, "message": msg, "detail": detail or []})

    def error(self, *a, **k):
        self.add("ERROR", *a, **k)

    def warn(self, *a, **k):
        self.add("WARN", *a, **k)

    def info(self, *a, **k):
        self.add("INFO", *a, **k)

    @property
    def n_error(self):
        return sum(1 for x in self.items if x["level"] == "ERROR")


# ---------------------------------------------------------------- 個別の検査

def check_balance(rows, rep):
    """借貸一致。台帳全体と、対象月ごとの両方を見る。"""
    dr = sum(num(r.get("借方金額")) for r in rows)
    cr = sum(num(r.get("貸方金額")) for r in rows)
    if round(dr - cr, 2) != 0:
        rep.error("借貸一致", f"借方 {dr:,.0f} ≠ 貸方 {cr:,.0f}（差 {dr-cr:,.0f}）")
    else:
        rep.info("借貸一致", f"借方=貸方={dr:,.0f}")

    per = defaultdict(lambda: [0, 0])
    for r in rows:
        m = str(r.get("対象月") or "")[:7]
        per[m][0] += num(r.get("借方金額"))
        per[m][1] += num(r.get("貸方金額"))
    bad = [f"{m}: 差 {d-c:,.0f}" for m, (d, c) in sorted(per.items()) if round(d - c, 2) != 0]
    if bad:
        rep.error("借貸一致(月別)", f"{len(bad)}ヶ月で不一致", bad)


def check_journal_id(rows, rep):
    """仕訳IDの重複。出所グループごとに採番すると衝突する（実際に4件発生した）。"""
    seen = defaultdict(list)
    for r in rows:
        jid = str(r.get("仕訳ID") or "").strip()
        if jid:
            seen[jid].append(r["__row"])
    dup = {k: v for k, v in seen.items() if len(v) > 1}
    if dup:
        rep.error("仕訳ID重複", f"{len(dup)}件のIDが重複",
                  [f"{k} → 行 {', '.join(map(str, v))}" for k, v in sorted(dup.items())])
    else:
        rep.info("仕訳ID重複", f"なし（{len(seen)}件すべて一意）")


def check_ref_dup(rows, rep):
    """参照ID×出所×対象月 の重複＝同じ元データから2回仕訳を起こしている。

    キーに対象月を含めるのが肝。合算請求は対象月をまたぐことがあり（INV-2026-07-016 は
    2026-06 と 2026-07 の行を持つ）、参照ID×出所 の2キーだと正常な月別計上まで重複と誤検知する。
    そこを避けようとして参照IDに枝番（INV-xxx/C-029）を付けると、今度は重複防止そのものが
    効かなくなって二重計上を許す。正解は「枝番を作らず、キーに対象月を足す」。
    """
    seen = defaultdict(list)
    for r in rows:
        key = (str(r.get("参照ID") or "").strip(),
               str(r.get("出所") or "").strip(),
               str(r.get("対象月") or "").strip())
        if key[0]:
            seen[key].append(r)
    # 取消仕訳（借貸を反転させた行）とその元仕訳の組は重複ではない。
    # 借方科目・貸方科目の組が互いに逆で金額が一致する行を打ち消し合わせ、残りが2件以上なら重複。
    dup = {}
    for k, rs in seen.items():
        if len(rs) < 2:
            continue
        remaining = list(rs)
        i = 0
        while i < len(remaining):
            a = remaining[i]
            partner = next((b for b in remaining[i + 1:]
                            if b.get("借方科目") == a.get("貸方科目")
                            and b.get("貸方科目") == a.get("借方科目")
                            and num(b.get("借方金額")) == num(a.get("借方金額"))), None)
            if partner is not None:
                remaining.remove(partner)
                remaining.remove(a)
            else:
                i += 1
        if len(remaining) > 1:
            dup[k] = [r["__row"] for r in rs]
    if dup:
        rep.error("参照ID×出所×対象月 重複", f"{len(dup)}組が重複",
                  [f"{k[0]} / {k[1]} / {k[2]} → 行 {', '.join(map(str, v))}"
                   for k, v in sorted(dup.items())])
    else:
        rep.info("参照ID×出所×対象月 重複", "なし")


BRANCH_RE = re.compile(r"^(INV-[0-9]{4}-[0-9]{2}-[0-9]+)\s*[/／]")


def check_ref_branch(rows, rep):
    """参照IDの枝番（INV-xxx/C-029）を禁止する。

    枝番を付けると同じ請求書が別の参照IDに化けて、上の「参照ID×出所」重複防止をすり抜ける。
    NITOH で 3,575,000 を売上・入金とも二重計上した実際の経路がこれ。
    """
    bad = []
    for r in rows:
        ref = str(r.get("参照ID") or "").strip()
        m = BRANCH_RE.match(ref)
        if m:
            bad.append(f"行 {r['__row']}: {ref} → 正: {m.group(1)}（内訳は摘要へ）")
    if bad:
        rep.error("参照IDの枝番", f"{len(bad)}件。重複防止をすり抜けます", bad)
    else:
        rep.info("参照IDの枝番", "なし")


def check_accounts(rows, master, rep):
    known = {str(r.get("科目名") or "").strip() for r in master if r.get("科目名")}
    used = set()
    for r in rows:
        for k in ("借方科目", "貸方科目"):
            v = str(r.get(k) or "").strip()
            if v:
                used.add(v)
    unknown = sorted(used - known)
    if unknown:
        rep.error("勘定科目マスタ外", f"{len(unknown)}科目がマスタに無い", unknown)
    else:
        rep.info("勘定科目マスタ外", f"なし（{len(used)}科目を使用）")

    # MF科目名 が空の科目は、会計SaaSへ送るときに変換できず素通りする
    nomap = sorted(str(r.get("科目名")).strip() for r in master
                   if r.get("科目名") and not str(r.get("MF科目名") or "").strip())
    if nomap:
        rep.warn("MF科目名 未設定", f"{len(nomap)}科目。会計SaaSへ送るとローカル名のまま飛びます", nomap)


def check_zero_rows(rows, rep):
    bad = [f"行 {r['__row']}: {r.get('摘要') or ''}"
           for r in rows if num(r.get("借方金額")) == 0 and num(r.get("貸方金額")) == 0]
    if bad:
        rep.warn("0円の仕訳", f"{len(bad)}件。0円の行は仕訳にしない", bad[:20])


def check_blank_rows(path, rep):
    """途中の空行の位置を報告する。

    「ヘッダ + 件数 + 1」で書き込み開始行を計算すると既存行を潰す。実際に11行潰した事故がある。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["仕訳帳"]
    blanks, last = [], 0
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if r and r[0] not in (None, ""):
            last = i
        elif i > 1:
            blanks.append(i)
    wb.close()
    blanks = [b for b in blanks if b < last]
    if blanks:
        rep.warn("途中の空行", f"{len(blanks)}行。件数から開始行を計算しないこと（実最終行={last}）",
                 [str(b) for b in blanks])


def check_invoice_overbooking(jrows, arrows, rep):
    """請求書ID単位で、計上額が額面を超えていないか。二重計上の最後の砦。

    行数では判定しない（合算請求は同じ番号が複数行に付くのが正常）。額面と突き合わせる。
    """
    face = defaultdict(float)
    for r in arrows:
        inv = str(r.get("請求書番号") or "").strip()
        if inv and not inv.startswith("("):
            face[inv] += num(r.get("請求額(税込)"))

    booked_sales = defaultdict(float)
    booked_cash = defaultdict(float)
    for r in jrows:
        ref = str(r.get("参照ID") or "").strip()
        inv = ref.split("/")[0].split("／")[0].strip()
        if not inv.startswith("INV-"):
            continue
        src = str(r.get("出所") or "").strip()
        # 取消仕訳（借方=売上高／借方=売掛金 の反転行）は差し引く
        if src == "AR請求":
            if str(r.get("借方科目") or "").strip() == "売掛金":
                booked_sales[inv] += num(r.get("借方金額"))
            else:
                booked_sales[inv] -= num(r.get("借方金額"))
        elif src == "AR入金":
            if str(r.get("借方科目") or "").strip() == "普通預金":
                booked_cash[inv] += num(r.get("借方金額"))
            else:
                booked_cash[inv] -= num(r.get("借方金額"))

    over_s, over_c = [], []
    for inv, amt in sorted(booked_sales.items()):
        f = face.get(inv)
        if f is None:
            continue
        if round(amt - f, 2) > 0:
            over_s.append(f"{inv}: 計上 {amt:,.0f} > 額面 {f:,.0f}（超過 {amt-f:,.0f}）")
    for inv, amt in sorted(booked_cash.items()):
        f = face.get(inv)
        if f is None:
            continue
        if round(amt - f, 2) > 0:
            over_c.append(f"{inv}: 消込 {amt:,.0f} > 額面 {f:,.0f}（超過 {amt-f:,.0f}）")

    if over_s:
        rep.error("売上の二重計上", f"{len(over_s)}件が額面超過", over_s)
    else:
        rep.info("売上の二重計上", f"なし（{len(booked_sales)}請求書を検査）")
    if over_c:
        rep.error("入金の二重消込", f"{len(over_c)}件が額面超過", over_c)
    else:
        rep.info("入金の二重消込", f"なし（{len(booked_cash)}請求書を検査）")

    unknown = sorted(set(booked_sales) - set(face))
    if unknown:
        rep.warn("台帳に無い請求書ID", f"{len(unknown)}件。仕訳にあるが月次請求スケジュールに無い", unknown)


def check_ar_closure(jrows, rep):
    """売掛金の自閉。期首 + 売上 − 入金 = 残高。期首残高が無いとマイナスに沈む。"""
    opening = has_opening = 0
    dr = cr = 0
    for r in jrows:
        if str(r.get("出所") or "").strip() == "期首残高":
            has_opening += 1
            if str(r.get("借方科目") or "").strip() == "売掛金":
                opening += num(r.get("借方金額"))
        if str(r.get("借方科目") or "").strip() == "売掛金":
            dr += num(r.get("借方金額"))
        if str(r.get("貸方科目") or "").strip() == "売掛金":
            cr += num(r.get("貸方金額"))
    bal = dr - cr
    if not has_opening:
        rep.warn("期首残高", "出所=期首残高 の行がありません。BSが成立せず売掛金がマイナスに沈みます")
    if bal < 0:
        rep.error("売掛金残高", f"マイナス {bal:,.0f}（借方 {dr:,.0f} − 貸方 {cr:,.0f}）。"
                               "期首残高の欠落か、売上未計上のまま入金だけ計上している可能性")
    else:
        rep.info("売掛金残高", f"{bal:,.0f}（うち期首 {opening:,.0f}）")


def check_tax_basis(jrows, summary, rep):
    """税基準ズレの疑いを、比率から機械的に検出する。

    ローカル仕訳台帳は税込、会計SaaSのレポートは既定で税抜。揃えずに比べると約1.1倍ずれる。
    月次サマリ（MF取込値）と仕訳帳の集計を月ごとに比べ、比率が 1.10 / 0.909 に近ければ警告する。
    実際にこれを見落として、丸一セッション誤った差額を追いかけた。
    """
    local = defaultdict(float)
    for r in jrows:
        if str(r.get("貸方科目") or "").strip() == "売上高":
            local[str(r.get("対象月") or str(r.get("日付") or ""))[:7]] += num(r.get("貸方金額"))
        if str(r.get("借方科目") or "").strip() == "売上高":
            local[str(r.get("対象月") or str(r.get("日付") or ""))[:7]] -= num(r.get("借方金額"))

    hits = []
    for r in summary:
        m = str(r.get("対象月") or "")[:7]
        mf = num(r.get("売上高"))
        lo = local.get(m, 0)
        if mf <= 0 or lo <= 0:
            continue
        ratio = lo / mf
        if 1.08 <= ratio <= 1.12 or 0.89 <= ratio <= 0.93:
            hits.append(f"{m}: ローカル {lo:,.0f} / サマリ {mf:,.0f} = 比 {ratio:.3f}")
    if hits:
        rep.error("税基準ズレの疑い", "差が約1.1倍です。MFは include_tax: true で取り直してください", hits)
    elif summary:
        rep.info("税基準", "1.1倍の疑いなし")


# ---------------------------------------------------------------- 実行

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("ledger_dir", help="台帳フォルダ（仕訳台帳.xlsx などがある場所）")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    d = Path(args.ledger_dir)
    rep = Report()

    jpath = d / JOURNAL
    if not jpath.exists():
        sys.exit(f"{jpath} がありません")

    _, jrows = read_tab(jpath, "仕訳帳")
    jrows = [r for r in jrows if str(r.get("仕訳ID") or "").strip()]
    _, master = read_tab(jpath, "勘定科目マスタ")
    try:
        _, summary = read_tab(jpath, "月次サマリ")
    except KeyError:
        summary = []

    arrows = []
    if (d / AR).exists():
        _, arrows = read_tab(d / AR, "月次請求スケジュール")

    check_balance(jrows, rep)
    check_journal_id(jrows, rep)
    check_ref_dup(jrows, rep)
    check_ref_branch(jrows, rep)
    check_accounts(jrows, master, rep)
    check_zero_rows(jrows, rep)
    check_blank_rows(jpath, rep)
    if arrows:
        check_invoice_overbooking(jrows, arrows, rep)
    check_ar_closure(jrows, rep)
    check_tax_basis(jrows, summary, rep)

    if args.json:
        print(json.dumps({"errors": rep.n_error, "items": rep.items}, ensure_ascii=False, indent=2))
    else:
        icon = {"ERROR": "✗", "WARN": "!", "INFO": "✓"}
        for it in rep.items:
            print(f"{icon[it['level']]} {it['check']}: {it['message']}")
            for x in it["detail"][:30]:
                print(f"    {x}")
            if len(it["detail"]) > 30:
                print(f"    …他 {len(it['detail'])-30} 件")
        print()
        print(f"仕訳 {len(jrows)}行 / ERROR {rep.n_error}件")

    return 1 if rep.n_error else 0


if __name__ == "__main__":
    sys.exit(main())
